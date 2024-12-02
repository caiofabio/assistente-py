import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

from config import PATH_FICHA, USER_FICHA

# Caminho para o seu arquivo Excel
arquivo_excel = PATH_FICHA

# Abra o arquivo Excel
workbook = openpyxl.load_workbook(arquivo_excel)
sheet = workbook.active

# Funcao para inserir uma data em uma celula especifica, lidando com celulas mescladas
def inserir_data(linha, coluna, data):
    cell = sheet.cell(row=linha, column=coluna)
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                sheet.unmerge_cells(str(merged_range))
                main_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                main_cell.value = data
                sheet.merge_cells(str(merged_range))
                return
    else:
        cell.value = data

# Funcao para inserir texto em uma celula especifica
def inserir_texto(linha, coluna, texto):
    cell = sheet.cell(row=linha, column=coluna)
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                sheet.unmerge_cells(str(merged_range))
                main_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                main_cell.value = texto
                sheet.merge_cells(str(merged_range))
                return
    else:
        cell.value = texto


# Obter a data de hoje
hoje = datetime.now()

inserir_texto(5, 2, USER_FICHA)  # Coluna C e a terceira coluna
inserir_texto(5, 13, f"{hoje.month}/{hoje.year}")  # Coluna C e a terceira coluna

# Calcular a data da ultima segunda-feira
if hoje.weekday() == 6:  # Se hoje for domingo
    ultima_segunda = hoje - timedelta(days=6)
else:
    ultima_segunda = hoje - timedelta(days=hoje.weekday())

# Preencher datas de B12 ate B28 e texto "casa" em C12 ate C28
linha_inicial = 12
linha_final = 28

for i in range(linha_inicial, linha_final + 1):
    data_para_inserir = ultima_segunda + timedelta(days=(i - linha_inicial))
    if data_para_inserir > hoje:
        break
    inserir_data(i, 2, data_para_inserir)  # Coluna B e a segunda coluna
    inserir_texto(i, 3, "CASA")  # Coluna C e a terceira coluna
    inserir_texto(i, 4, "08:00")  # Coluna C e a terceira coluna
    inserir_texto(i, 5, "12:00")  # Coluna C e a terceira coluna
    inserir_texto(i, 6, "13:00")  # Coluna C e a terceira coluna
    inserir_texto(i, 7, "17:00")  # Coluna C e a terceira coluna

# Salve o arquivo Excel modificado
workbook.save(arquivo_excel)
